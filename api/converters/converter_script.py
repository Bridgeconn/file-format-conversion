import csv
import re
import sys
import glob
import errno
import os
import pathlib
from bs4 import BeautifulSoup
import openpyxl

def usfm_to_csv(file):

  file_base_path, file_name = os.path.split(file)

  file_root_path = pathlib.Path().absolute()
  print(file_root_path)

  root_path = str(file_root_path)+'/'+file
  print(root_path)
  target_file_path = root_path.split('.')[0] + '.csv'


  # open usfm file for reading
  f = open (root_path,'r')

  # open csv file for writing
  outfile = open(target_file_path, 'w')

  # create the csv writer object
  csvwriter = csv.writer(outfile, delimiter ='\t')

  # Writing to csv file
  prev_book = ""
  prev_chapter = ""
  chapter = ""
  book = ""
  verse = ""
  addline = ""

  d = f.readlines()
  f.seek(0)
  for line in d:
    if line == "\n":
        if addline:
            csvwriter.writerow([prev_book, prev_chapter, verse, addline])
            addline = ""
        continue
    elif line[0:3] == '\c ':
      if addline:
        csvwriter.writerow([prev_book, prev_chapter, verse, addline])
        addline = ""
      chapter = line[3:5]
      if chapter == prev_chapter:
        continue
      else:
        prev_chapter = chapter
    elif line[0:4] == '\id ':
      if addline:
        csvwriter.writerow([prev_book, prev_chapter, verse, addline])
        addline = ""
      book = line[4:].strip()
      if book == prev_book:
        continue
      else:
        prev_book = book
    elif line[0:3] == '\\v ':
      if addline:
        csvwriter.writerow([prev_book, prev_chapter, verse, addline])
        addline = ""
        verse = line[3:5]
        addline = line[5:]
    elif line[0:4] == '\q2 ' or line[0:4] == '\q3 ':
      if line[4:] != " ":
        addline = addline.strip() + " " + line[4:].strip()
    elif line[0:3] == '\q ' or line[0:3] == '\m ':
      if line[4:] != " ":
        addline = addline.strip() + " " + line[3:].strip()

    if (line == d[-1]):
      csvwriter.writerow([prev_book, prev_chapter, verse, addline])

    prev_book = book
    prev_chapter = chapter
  
  print(target_file_path)
  return target_file_path

  f.close()
  outfile.close()

# Tested
# t1 = usfm_to_csv("uploads/18199276-08aa-434d-805e-5edd871a5bd6/GEN.usfm")
# print(t1)


def csv_to_usfm(file):
  idCount=0
  file_base_path, file_name = os.path.split(file)

  file_root_path = pathlib.Path().absolute()
  print(file_root_path)

  root_path = str(file_root_path)+'/'+file
  print(root_path)
  target_file_path = root_path.split('.')[0] + '.usfm'



  # open csv file for reading
  f = open (root_path,'r')

  # open usfm file for writing
  outfile = open(target_file_path, 'w')
  reader = csv.reader(f, delimiter='\t')
  for row in reader:
    outfile.write("\n\id " + row[0] + "\n\c " + row[1] + "\n\\v " + row[2] + " " +row[3])
    # outfile.close()

    prev_book = ""
    prev_chapter = ""
    chapter = ""
    book = ""

    # Open csv file again for removal of all extra tags
    f = open (root_path,'r+')
    d = f.readlines()
    f.seek(0)
    for line in d:
        if line == "\n":
            continue
        elif line[0:3] == '\c ':
            chapter = line[3]
            if chapter == prev_chapter:
                continue
            else:
                f.write(line)
        elif line[0:4] == '\id ':
            book = line[4]
            if book == prev_book:
                continue
            else:
                f.write(line)
        else:
            f.write(line)
        prev_book = book
        prev_chapter = chapter

    # Close the files
    f.truncate()
    return target_file_path
    f.close()
   
    print ("Conversion from csv to usfm done !")


#Tested
# t2 = csv_to_usfm("uploads/1JN.csv")
# print (t2)

def md_to_csv(file):
  file_base_path, file_name = os.path.split(file)

  file_root_path = pathlib.Path().absolute()
  print(file_root_path)

  root_path = str(file_root_path)+'/'+file
  print(root_path)
  target_file_path = root_path.split('.')[0] + '.csv'


  # open usfm file for reading
  # f = open (root_path,'r')

  # open csv file for writing
  outfile = open(target_file_path, 'w')
  csv_file = 'outputcsv_file.csv'
  outfile = open(csv_file,'w')
  csvwriter = csv.writer(outfile)
  csvwriter.writerow([ "id" ,"keyword", "wordforms","strongs","definition", "translationhelp","seealso","ref","examples"])
  idCount=0
  f = open (root_path,'r')
  d = f.readlines()
  f.seek(0)
  wordforms = d[0].replace('#','').strip()
  idCount+= 1
  # keyword=os.path.basename(name).replace(".md","")
  for line in d:
          if  "* Strong's:" in line:
                strongs= line.replace("* Strong","").strip()
                continue;
          if  "यह भी देखें:" in line:
                seealso= line.replace("(यह भी देखे:","").strip()
                continue

  i = 4
  definition=""
  translationhelp=""
  while i < len(d) :
          if "यह भी देखें" in d[i] or d[i].startswith("## बाइबल स"):
                  break
          elif d[i].startswith("## अनुवाद के सुझाव") or translationhelp!="":
                  ranslationhelp=translationhelp+d[i]
          else:
                  definition=definition+d[i]
          i+=1

  ref=""
  refContentIndex = 0;
  i=6
  while i < len(d):
          if d[i].startswith("## बाइबल स"):
                  refContentIndex=i+1
                  break
          i+=1
  while  refContentIndex >0 and refContentIndex < len(d):
          if d[refContentIndex].startswith("## शब्द तथ्य") or d[refContentIndex].startswith("## बाइबल कहानियों से उदाहरण") or d[refContentIndex].startswith("* Strong") :
                  break
          else:
                  ref = ref + d[refContentIndex]
          refContentIndex += 1

  examples=""
  examplesContentIndex = 0;
  i = 9
  while i < len(d):
          if d[i].startswith("## बाइबल कहानियों से उदाहरण"):
                  examplesContentIndex = i + 1
                  break
          i += 1
  while examplesContentIndex > 0 and examplesContentIndex < len(d):
          if d[examplesContentIndex].startswith("## शब्द तथ्य") or d[examplesContentIndex].startswith("* Strong"):
                  break
          else:
                  examples = examples + d[examplesContentIndex]
          examplesContentIndex += 1
  csvwriter.writerow([idCount,wordforms,strongs,definition,translationhelp,seealso,ref,examples])
  return target_file_path
  f.close()
  outfile.close()

#Tested
# t3 = md_to_csv("uploads/aaron.md")
# print (t3)

def html_to_csv(file):
  file_base_path, file_name = os.path.split(file)

  file_root_path = pathlib.Path().absolute()
  print(file_root_path)

  root_path = str(file_root_path)+'/'+file
  print(root_path)
  target_file_path = root_path.split('.')[0] + '.csv'


  # open usfm file for reading
  f = open (root_path,'r')

  # open csv file for writing
  outfile = open(target_file_path, 'w')
  soup = BeautifulSoup(open(root_path,'r'), "lxml")

  outfile = open(target_file_path, 'w')

  # Create the csv writer object
  csvwriter = csv.DictWriter(outfile, fieldnames = ["Chapter", "Verse", "Eng Word"], delimiter ='\t')
  csvwriter.writeheader()
  # create the csv writer object
  csvwriter = csv.writer(outfile, delimiter ='\t')
  print ("Parsing HTML file")

  for t in soup.find_all('div', id = re.compile("test_")):
      s = t['id'].split('_')
      s = s[2]
      s = s.split('-')
      chapter = s[0]
      verse = s[1]
      en = t.get_text().strip()
      csvwriter.writerow([chapter, verse, en])


  print ("Converted to CSV !")
  return target_file_path
  outfile.close()
  
# #Tested
# t4 = html_to_csv("uploads/chapter1.html")
# print (t4)

def xlsx_to_txt(file):
        # Open xlsx file
        file_base_path, file_name = os.path.split(file)

        file_root_path = pathlib.Path().absolute()
        print(file_root_path)

        root_path = str(file_root_path)+'/'+file
        print(root_path)
        target_file_path = root_path.split('.')[0] + '.txt'
        wb = openpyxl.load_workbook(file)
        sheet = wb.get_sheet_by_name('Sheet1')

        f = open (root_path,'r')

        # open txt file for writing
        outfile = open(target_file_path, 'w')

        # Access each element of a column row by row and write it to txt file
        for row in range(1, sheet.max_row + 1):
                name = sheet['C' + str(row)].value
                outfile.write(str(name))
                outfile.write("\n")

        # Close the file and print result
        return target_file_path
        outfile.close()
        print ('Done.')

#Tested
# t5 = xlsx_to_txt("uploads/example.xlsx")
# print (t4)

# def xlsx_to_csv(Source_files):
# #     for s_file in Source_files:
# bookName = Source_files.split("/")[-1].split(".")[0]
# fileName = Source_files.split("/")[-1]
# # s_Path = glob.glob(os.getcwd() + "/hi_tn/s_tsv/" + str(fileName))
# s_Path =str(fileName)
# # t_Path = glob.glob(os.getcwd() + "/hi_tn/xlsx/" + str(bookName) + ".xlsx")
# t_Path = (str(bookName) + ".xlsx")
# print(s_Path)
# print(t_Path)
# print(bookName)
# # Open Target File
# wb_obj = openpyxl.load_workbook(t_Path[0])
# sheet_obj = wb_obj.active
# max_col = sheet_obj.max_column
# max_row = sheet_obj.max_row

# eng_row_count = 0
# hindi_row_count = 1


# label_file = (bookName + ".tsv")
# with open(label_file, 'w', encoding='utf-8') as tsv_file:
#         twriter = csv.writer(tsv_file, delimiter='\t')
#         with open(s_Path[0], 'r', encoding='utf-8') as tsvfile:
#         reader = csv.reader(tsvfile, delimiter='\t')
#         for rows in range(max_row):
#                 try:
#                 row = next(reader)
#                 except:
#                 break
#                 if row[8] == '':
#                 pass
#                 else:
#                 # dic1 = []
#                 combine_cells = ''
#                 eng_ocr_note = row[8]
#                 split_eng_ocr = eng_ocr_note.split('\n')
#                 find_array_len = len(split_eng_ocr)
#                 if (len(split_eng_ocr) > 1):
#                         subline_count = 0
#                         array_count = find_array_len
#                         for sub_lines in split_eng_ocr:
#                         split_sublines = sub_lines.split('\t')
#                         if sub_lines == '':
#                                 pass
#                         elif subline_count == 0:
#                                 trans_ocr_note = sheet_obj.cell(row=hindi_row_count, column=4).value
#                                 hin_ocr_note0 = sheet_obj.cell(row=hindi_row_count, column=5).value
#                                 hin_ocr_note = str(hin_ocr_note0).strip()
#                                 combine_cells += hin_ocr_note + '\n'
#                                 subline_count += subline_count + 1
#                                 hindi_row_count += 1
#                         elif array_count == 1:
#                                 hin_book = sheet_obj.cell(row=hindi_row_count, column=1).value
#                                 hin_book1 = hin_book.split("\n")[0]
#                                 hin_chaptr = sheet_obj.cell(row=hindi_row_count, column=2).value
#                                 hin_verse = sheet_obj.cell(row=hindi_row_count, column=3).value
#                                 trans_ocr_note = sheet_obj.cell(row=hindi_row_count, column=4).value
#                                 hin_ocr_note0 = sheet_obj.cell(row=hindi_row_count, column=5).value
#                                 hin_ocr_note = str(hin_ocr_note0).strip()
#                                 combine_cells += str(hin_book1).strip() + "\t" + str(hin_chaptr).strip() + "\t" + str(hin_verse).strip() + "\t" + str(split_sublines[3]) + "\t" + str(split_sublines[4]) + "\t" + str(split_sublines[5]) + "\t" + str(split_sublines[6]) + "\t" + str(split_sublines[7]) + "\t" + str(hin_ocr_note)
#                                 hindi_row_count += 1
#                         else:
#                                 hin_book = sheet_obj.cell(row=hindi_row_count, column=1).value
#                                 hin_book1 = hin_book.split("\n")[0]
#                                 hin_chaptr = sheet_obj.cell(row=hindi_row_count, column=2).value
#                                 hin_verse = sheet_obj.cell(row=hindi_row_count, column=3).value
#                                 trans_ocr_note = sheet_obj.cell(row=hindi_row_count, column=4).value
#                                 hin_ocr_note0 = sheet_obj.cell(row=hindi_row_count, column=5).value
#                                 hin_ocr_note = hin_ocr_note0.strip()
#                                 combine_cells += str(hin_book1).strip() + "\t" + str(hin_chaptr).strip() + "\t" + str(hin_verse).strip() + "\t" + str(split_sublines[3]) + "\t" + str(split_sublines[4]) + "\t" + str(split_sublines[5]) + "\t" + str(split_sublines[6]) + "\t" + str(split_sublines[7]) + "\t" + str(hin_ocr_note) +'\n'
#                                 hindi_row_count += 1
#                         array_count -= 1
#                         find_link_source = re.findall(r'(\[\[\w+\:[\/\w+\-]*\]\])', eng_ocr_note)
#                         find_link_target = re.findall(r'@', combine_cells)
#                         edited_targetl = combine_cells
#                         dic1 = []
#                         if find_link_source:
#                         for link in find_link_source:
#                                 if link == '':
#                                 pass
#                                 else:
#                                 dic1.append(link)
#                         for k in dic1:
#                                 edited_targetl = edited_targetl.replace("@", k, 1)
#                         rep_br = re.sub(r'\$', '<br>', edited_targetl)
#                         twriter.writerow([str(row[0]), str(row[1]), str(row[2]), str(row[3]), str(row[4]), str(row[5]), str(row[6]), str(row[7]), str(rep_br).strip()])        

#                         print(eng_ocr_note)
#                         print("------------------------------------------------------")
#                         print(str(rep_br.strip()))
#                         print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>") 
#                         print(" ")
#                         print(" ")

#                         else:
#                         rep_br = re.sub(r'\$', '<br>', combine_cells)
#                         twriter.writerow([str(row[0]), str(row[1]), str(row[2]), str(row[3]), str(row[4]), str(row[5]), str(row[6]), str(row[7]), str(rep_br).strip()])

#                         print(eng_ocr_note)
#                         print("------------------------------------------------------")
#                         print(str(rep_br.strip()))
#                         print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>") 
#                         print(" ")
#                         print(" ")

#                 else:
#                         find_link_source = re.findall(r'(\[\[\w+\:[\/\w+\-]*\]\])', eng_ocr_note)
#                         trans_ocr_note = sheet_obj.cell(row=hindi_row_count, column=4).value
#                         hin_ocr_note0 = sheet_obj.cell(row=hindi_row_count, column=5).value
#                         hin_ocr_note = str(hin_ocr_note0).strip()
#                         if hin_ocr_note == None:
#                         break
#                         find_link_target = re.findall(r'@',hin_ocr_note)
#                         edited_targetl = hin_ocr_note 
#                         hindi_row_count += 1
#                         dic1 = []

#                         if find_link_source:
#                         for link in find_link_source:
#                                 if link == '':
#                                 pass
#                                 else:
#                                 dic1.append(link)   
#                         for k in dic1:
#                                 edited_targetl = edited_targetl.replace("@", k, 1)
#                         rep_br = re.sub(r'\$', '<br>', edited_targetl)
#                         twriter.writerow([str(row[0]), str(row[1]), str(row[2]), str(row[3]), str(row[4]), str(row[5]), str(row[6]), str(row[7]), str(rep_br).strip()]) 
#                         print(eng_ocr_note)
#                         print("------------------------------------------------------")
#                         print(str(rep_br.strip()))
#                         print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>") 
#                         print(" ")
#                         print(" ")

#                         else:
#                         rep_br = re.sub(r'\$', '<br>', hin_ocr_note)
#                         twriter.writerow([str(row[0]), str(row[1]), str(row[2]), str(row[3]), str(row[4]), str(row[5]), str(row[6]), str(row[7]), str(rep_br).strip()])
#                         print(eng_ocr_note)
#                         print("------------------------------------------------------")
#                         print(str(rep_br.strip()))
#                         print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>") 
#                         print(" ")
#                         print(" ")
# t5 = xlsx_to_csv("uploads/en_tn_57-TIT.xlsx")
# print (t5)