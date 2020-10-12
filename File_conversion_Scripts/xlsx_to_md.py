
# -*- coding: utf-8 -*-

import openpyxl
import csv
import re
import glob
import os


# files = glob.glob(os.getcwd() + "/Source/*.xlsx")
# print(files)
def xlsx_to_md(files_xlsx):
    for fl in files_xlsx:
        bookname = fl.split("/")[-1].split('.')[0]
        print(bookname)
        wb_obj = openpyxl.load_workbook(fl)
        sheet_obj = wb_obj.active
        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row
        
        # Creating folder 
        os.makedirs(bookname)
        s_path = glob.glob(os.getcwd()+'/'+bookname +'/')
        
        # print(s_path)
        folders = 'Chapter'
        for i in range(2, max_row + 1):
            book = sheet_obj.cell(row=i, column=1).value
            chapter = sheet_obj.cell(row=i, column=2).value
            if chapter == folders:
                pass
            else:
                # print(chapter.zfill(2))
                os.makedirs(s_path[0] + chapter.zfill(2))
                folders = chapter
            verse = sheet_obj.cell(row=i, column=3).value
            split_verse = verse.split('-')[0]
            question = sheet_obj.cell(row=i, column=5).value
            # print(question)
            with open(s_path[0] + folders.zfill(2) + '/' + split_verse.zfill(2) + '.md', "w+") as file:
                try:
                    for q in question.split('\n'):
                        q_match = re.match(r'ಪ್ರ\? ?',q)
                        a_match = re.match(r'ಉ\. ?',q)
                        if q_match:
                            q_sub = re.sub(r'ಪ್ರ\? ?','# ',q + '\n')
                            file.write(q_sub)
                            # print(q_sub)
                        elif a_match:
                            a_sub = re.sub(r'ಉ\. ?','',q + '\n\n')
                            file.write(a_sub)
                            # print((a_sub))
                        else:
                            pass
                except:
                    # pass
                    with open(s_path[0] + folders.zfill(2) + '/' + split_verse.zfill(2) + '.md', "w+") as file:
                        file.write("")
        
    




































#   -------------------------------------------------------------
  
  
  
  
  
    # for row in sheet_obj.iter_rows (min_row=1, min_col=1, max_row =sheet_obj.max_row, max_col=3):
    #     for cell in row:
    #         print(cell.value)
    #         # print()
    #     # # print('--------------------------------')
    #     # print(hin_ocr_note)
    #     # print('--------------------------------')











#     source_content = []
#     filename = fl.split("/")[-1]
#     s_path = glob.glob(os.getcwd() + "/source/" + str(filename))
#     t_path = glob.glob(os.getcwd() + "/target/" + str(bookname) + ".xlsx")

#     # ----------- Target File ----------------#
#     wb_obj = openpyxl.load_workbook(t_path[0])
#     sheet_obj = wb_obj.active
#     max_col = sheet_obj.max_column
#     max_row = sheet_obj.max_row
#     row_count = 1

#     # ---------- Source File -----------------#
#     label_file = (bookname + ".tsv")
#     # print (label_file)
#     with open(label_file, 'w', encoding = 'utf-8') as tsv_file:
#         twriter = csv.writer(tsv_file, delimiter='\t')
#         with open(s_path[0], 'r', encoding = 'utf-8') as tsvfile:
#             reader = csv.reader(tsvfile, delimiter='\t')
#             for rows in reader:
#                 dic = []
#                 if rows[8] == '':
#                     print("english content (row) is empty")
#                 else:
#                     eng_ocr_note = rows[8]  
#                     hin_ocr_note = sheet_obj.cell(row=row_count, column=4).value
#                     if hin_ocr_note == None:
#                         print("hindi content (row) is empty")
#                     else:
#                         find_hind_see = re.findall(r'\( ?देखें:.*?\)', hin_ocr_note)
                        
#                         find_see = re.findall(r'\(See:.*?\]?\]?\)', eng_ocr_note)                  
#                         rep_see = re.sub(r'\(देखें:.*?\)', "देखें:", hin_ocr_note)
#                         # print(find_see)
#                         row_count += 1
#                         # print("=======")
#                         if find_see:
#                             for see in find_see:
#                                 if see == '':
#                                     pass
#                                 else:
#                                     dic.append(see)
#                             for k in dic:
#                                 rep_see = rep_see.replace("देखें:", k, 1)
#                             rep_see1 = re.sub(r'See', 'देखें', rep_see)
#                             rep_and = re.sub(r'and', 'और', rep_see1)
#                             rep_br = re.sub(r'\$', '<br>', rep_and)
#                             twriter.writerow([rows[0], rows[1], rows[2], rows[3], rows[4], rows[5], rows[6], rows[7], rep_br])
#                             # print("=============")
#                             # print(rep_br)

#                         else:
#                             rep_see_br = re.sub(r'\$', '<br>', hin_ocr_note)
#                             twriter.writerow([rows[0], rows[1], rows[2], rows[3], rows[4], rows[5], rows[6], rows[7], rep_see_br])
#                             # print(rep_see_br)
            
#             print("Completed")















# if rows[8] == '':
#                     print("english content (row) is empty")
#                     break
#                 else:
#                     eng_ocr_note = rows[8]  
#                     hin_ocr_note = sheet_obj.cell(row=row_count, column=4).value
#                     # find_see = re.findall(r'\(See: ?', eng_ocr_note)                  
#                     # eng_see += len(find_see)
#                     # print(len(find_see))
#                 if hin_ocr_note == None:
#                     print("hindi content (row) is empty")
#                     break
#                 else:
#                     # find_hind_see = re.findall(r'\( ?देखें:.*?\)', hin_ocr_note)
#                     # hin_see += len(find_hind_see)
#                     rep_see = re.sub(r'\(देखें:.*?\)', "देखें:", hin_ocr_note)
#             # print(eng_see)
#             # print(hin_see)
#                     find_see = re.findall(r'\(See: ?', eng_ocr_note)
#                     if find_see:
#                         for see in find_see:
#                             if see == '':
#                                 pass
#                             else:
#                                 dic[see] = "देखें:"
#                         for k, v in dic.items():
#                             rep_see = rep_see.replace(v, k, 1)
#                         rep_see1 = re.sub(r'See', 'देखें', rep_see)
#                         rep_and = re.sub(r'and', 'और', rep_see1)
#                         rep_br = re.sub(r'$', '<br>', rep_and)
#                         print(rep_br)
#                         # twriter.writerow([rows[0], rows[1], rows[2], rows[3], rows[4], rows[5], rows[6], rows[7], rep_br])
#                         row_count += 1
#                     else:
#                         # print(rep_see)
#                         # twriter.writerow([rows[0], rows[1], rows[2], rows[3], rows[4], rows[5], rows[6], rows[7], rep_see])
#                         row_count += 1
#             print("Completed")